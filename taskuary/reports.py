"""Scheduled report connections: pull from the systems you already have, drop informational
rows on the timeline (never tasks). A report source = a source row with Channel='report'
and ConfigJson {"type", "title", "every_minutes"/"daily_at", ...executor keys}.

REGISTRY: type -> executor(config) -> (headline, summary). Implemented: sqlite, rest, rss
(mssql with the [mssql] extra). Planned types fail loudly so a misconfig is visible on the
timeline instead of silently absent. Adding a type = one ~15-line function + a REGISTRY
entry - PRs welcome.
"""
import io, json, re, sqlite3
from datetime import datetime, timedelta
from loguru import logger

PLANNED = ['sharepoint_list', 'google_sheets', 'graphql', 'smb_file',
           # systems of record. Intacct is BUILT (see run_intacct); the rest are named because
           # the category is the question people arrive with - "does this reach our ERP / our
           # EMR" - and an empty Corporate systems group answers that worse than a list does.
           'netsuite', 'quickbooks', 'sap', 'workday', 'adp',
           'epic', 'cerner', 'pointclickcare']   # smb_file is a NETWORK
# share and still planned; a path on this machine is local_file and works now

MAX_ROWS, BODY_CHARS, AI_CHARS = 200, 20000, 12000     # per report; override with cfg['max_rows']
SUMMARY_TOKENS = 1500     # a report summary is prose, not a triage verdict - give it room


def row_limit(cfg):
    """(limit, is it YOURS). The default is a safety net, not a number anybody chose - and
    the headline has to say which one it was, or "capped at 200" points the owner at a
    setting they never made and cannot find."""
    n = cfg.get('max_rows')
    return (max(1, int(n)), True) if n else (MAX_ROWS, False)


def rows_out(rows, limit, unit='rows', mine=True):
    """(headline, body) from executor rows, SAYING SO when the result was cut. A silent cap
    made the AI describe 20 rows of a TOP 500 query as 'all of them' - fetch one extra row
    and the headline can tell the truth instead."""
    more = len(rows) > limit
    rows = rows[:limit]
    why = (f'capped at {limit}' if mine else f'capped at the default {limit}') + ' — set "max rows" on this source to see more'
    head = f'{len(rows)} rows' + (f' ({why})' if more else '')
    return head.replace('rows', unit, 1), '\n'.join(json.dumps(r, default=str) for r in rows)[:BODY_CHARS]


def run_sqlite(cfg):
    """{"db": "path.db", "query": "SELECT ...", "max_rows": 200} - the local-first database report."""
    cx = sqlite3.connect(cfg['db']); cx.row_factory = sqlite3.Row
    lim, mine = row_limit(cfg)
    rows = [dict(r) for r in cx.execute(cfg['query']).fetchmany(lim + 1)]
    cx.close()
    return rows_out(rows, lim, mine=mine)


def run_mssql(cfg):
    """{"server", "database", "auth", "username", "password", "driver", "query", "max_rows"} -
    see mssql.py. Configure the connection entirely from the Connectors tab."""
    from .mssql import run_report
    return run_report(cfg)


def run_intacct(cfg):
    """{"object": "GLENTRY", "fields": [...], "filters": [["BATCH_DATE", ">=", "08/01/2026"]],
    "max_rows": 200} - one readByQuery against Sage Intacct. The five credentials live on the
    Intacct card; a report carries only what it is asking for.

    Leave "fields" out and every field on the object comes back, which is the right default for
    a list somebody wants to eyeball and the wrong one for GL detail - so say which columns you
    want when the object is wide."""
    from .intacct import query
    obj = (cfg.get('object') or '').strip()
    if not obj: raise RuntimeError('no Intacct object set - e.g. GLENTRY, APBILL, VENDOR, LOCATION')
    lim, mine = row_limit(cfg)
    rows = query(cfg, obj, cfg.get('fields'), cfg.get('filters'), limit=lim + 1, order=cfg.get('order'))
    return rows_out(rows, lim, mine=mine)


def run_intacct_fields(cfg):
    """{"object": "APBILL"} - what the object actually HAS in this company, custom fields and
    all. It is a report in its own right (schedule it and a new custom field shows up on the
    timeline), and it is what the composer reads before writing an Intacct report."""
    from .intacct import fields_of
    obj = (cfg.get('object') or '').strip()
    if not obj: raise RuntimeError('no Intacct object set')
    lim, mine = row_limit(cfg)
    return rows_out(fields_of(cfg, obj), lim, unit='fields', mine=mine)


AGENT_SYSTEM = ('You are running a SCHEDULED REPORT for a busy operator. Do exactly what the instruction '
                'says - use your tools, read what you need to read - and then answer with the report itself: '
                'plain text or markdown, concrete (numbers, names, dates, deltas), no preamble and no '
                'questions back. If something the instruction asks about cannot be found, say so in the report.')


def run_agent(cfg):
    """{"agent": "coder", "skill": "weekly-user-review", "prompt": "...", "cwd": "C:/repo", "model": "..."} -
    the AI itself as the source: a coding CLI agent (Connectors -> AI CLI agents) runs your saved
    SKILL (a slash command - "/weekly-user-review") and/or a prompt, on the schedule, and what it
    answers is the report. "cwd" is optional: a project-level skill lives in its repo, a user-level
    one runs from anywhere. The AI summary pass is usually unnecessary - the agent already wrote prose.

    This is the "run my Claude skill every Monday" report: the agent researches, reads the systems
    it has tools for, and files what it found onto the Timeline like any other report."""
    from .llm import make_cli_llm
    store = cfg.get('store')
    if store is None: raise RuntimeError('the agent source needs the store (run it through the reports pipeline)')
    skill, prompt = str(cfg.get('skill') or '').strip().lstrip('/'), str(cfg.get('prompt') or '').strip()
    if not skill and not prompt: raise RuntimeError('give the agent a skill (/name) or a prompt - or both')
    name = str(cfg.get('agent') or 'coder').strip()
    llm = make_cli_llm(store, name, cfg.get('model') or None, cwd=cfg.get('cwd') or None)
    if llm is None: raise RuntimeError(f'no CLI agent named {name!r} - add one under Connectors -> AI CLI agents')
    ask = (f'/{skill}' + (' ' if prompt else '') if skill else '') + prompt
    out = str(llm(AGENT_SYSTEM, ask) or '').strip()
    if not out: raise RuntimeError(f'{name} answered nothing')
    what = f'/{skill}' if skill else 'a prompt'
    return f'{name} ran {what} - {len(out.splitlines())} lines', out[:BODY_CHARS]


def run_rest(cfg):
    """{"url", "headers", "path": "a.b"} - GET a JSON endpoint, dot-path into it.

    Through webguard, because the same executor is reachable from POST /api/tools/run: the URL
    can come from an AGENT, whose context is full of mail this codebase calls data and never
    instructions. See webguard for what a fetch to 169.254.169.254 would otherwise be."""
    from . import webguard
    r = webguard.get(cfg['url'], headers=cfg.get('headers') or {})
    r.raise_for_status()
    data = r.json()
    for k in (cfg.get('path') or '').split('.'):
        if k: data = data[int(k)] if isinstance(data, list) else data.get(k)
    return (f'{len(data)} items' if isinstance(data, list) else 'ok'), json.dumps(data, indent=1, default=str)[:BODY_CHARS]


def run_rss(cfg):
    """{"url"} - latest titles from an RSS/Atom feed. Guarded like run_rest: same reason."""
    from . import webguard
    xml = webguard.get(cfg['url']).text
    titles = re.findall(r'<title>(?:<!\[CDATA\[)?(.*?)(?:\]\]>)?</title>', xml)[1:11]
    return f'{len(titles)} new items', '\n'.join(f'- {t}' for t in titles)[:4000]


def run_winrm(cfg):
    """{"host", "script"} - run PowerShell ON a remote Windows box (WinRM / PS remoting,
    your current Windows credentials) and report its output. A box you can RDP into is
    usually domain-joined and WinRM-reachable already; if not, run Enable-PSRemoting on
    it once (elevated)."""
    import subprocess
    host, script = cfg['host'], cfg['script']
    p = subprocess.run(['powershell', '-NoProfile', '-NonInteractive', '-Command',
                        f'Invoke-Command -ComputerName {host} -ScriptBlock {{ {script} }}'],
                       capture_output=True, text=True, encoding='utf-8', errors='replace', timeout=180)
    if p.returncode != 0: raise RuntimeError((p.stderr or p.stdout or 'remote run failed')[:500])
    out = (p.stdout or '').strip()
    return f'{len(out.splitlines())} lines from {host}', out[:BODY_CHARS]


def run_mcp(cfg):
    """{"cmd", "args", "tool", "tool_args"} - call any MCP server's tool. See mcp.py."""
    from .mcp import run_report
    return run_report(cfg)


def run_database(cfg):
    """{"query"} - ANY database by connection string (postgres/mysql/snowflake/... URLs via
    SQLAlchemy, raw ODBC strings via pyodbc). The string lives on the 'Any database'
    connector card; see db.py."""
    from .db import run_report
    return run_report(cfg)


def run_aws(cfg):
    """{"service", "operation", "params", "path"} - any boto3 call with the AWS card's keys."""
    from .aws import run_aws as _run
    return _run(cfg)


def run_s3(cfg):
    """{"bucket", "key" | "prefix"} - read an S3 object, or list under a prefix. See aws.py."""
    from .aws import run_s3_object
    return run_s3_object(cfg)


def run_cwlogs(cfg):
    """{"log_group", "pattern", "hours"} - grep a CloudWatch log group. See aws.py."""
    from .aws import run_cloudwatch_logs
    return run_cloudwatch_logs(cfg)


def run_azure(cfg):
    """{"path", "api_version"} - GET any Azure Resource Manager object. See azure.py."""
    from .azure import run_azure as _run
    return _run(cfg)


def run_azblob(cfg):
    """{"account", "container", "blob" | "prefix"} - read or list Azure blob storage."""
    from .azure import run_azure_blob
    return run_azure_blob(cfg)


def run_azlogs(cfg):
    """{"workspace_id", "query", "hours"} - KQL against a Log Analytics workspace."""
    from .azure import run_azure_logs
    return run_azure_logs(cfg)


def run_entra_users(cfg):
    """{"filter", "select"} - Entra ID people, over Graph on the Azure card's app. See azure.py."""
    from .azure import run_entra_users as _run
    return _run(cfg)


def run_entra_groups(cfg):
    """{"group"} - a group's transitive members, or every group when blank."""
    from .azure import run_entra_groups as _run
    return _run(cfg)


def run_entra_signins(cfg):
    """{"hours", "failed_only"} - Entra sign-in activity (needs P1/P2 + AuditLog.Read.All)."""
    from .azure import run_entra_signins as _run
    return _run(cfg)


def run_entra_licenses(cfg):
    """Licence SKUs with seats consumed vs spare - the unused-seat report."""
    from .azure import run_entra_licenses as _run
    return _run(cfg)


def run_prometheus(cfg):
    """{"query" (PromQL)} - an instant query; each series is a row of its labels + value.
    The base URL (and an optional bearer token) live on the Prometheus card."""
    import requests
    base = (cfg.get('base_url') or '').strip().rstrip('/')
    if not base: raise RuntimeError('no Prometheus base URL set - Connectors → Prometheus')
    hdr = {'Authorization': f"Bearer {cfg['token']}"} if cfg.get('token') else {}
    r = requests.get(f'{base}/api/v1/query', params={'query': cfg['query']}, headers=hdr, timeout=30)
    r.raise_for_status()
    j = r.json()
    if j.get('status') != 'success': raise RuntimeError(f"prometheus: {j.get('error') or j}")
    rows = [{**(s.get('metric') or {}), 'value': (s.get('value') or [None, None])[1]}
            for s in (j.get('data') or {}).get('result') or []]
    lim, mine = row_limit(cfg)
    return rows_out(rows, lim, unit='series', mine=mine)


def run_datadog(cfg):
    """{"name" (optional filter)} - your monitors and their states, the at-a-glance health
    board. Keys live on the Datadog card (api key write-only + application key)."""
    import requests
    site = (cfg.get('site') or 'datadoghq.com').strip()
    params = {'name': cfg['name']} if cfg.get('name') else {}
    r = requests.get(f'https://api.{site}/api/v1/monitor', params=params, timeout=30,
                     headers={'DD-API-KEY': cfg.get('api_key') or '', 'DD-APPLICATION-KEY': cfg.get('app_key') or ''})
    if r.status_code in (401, 403): raise RuntimeError(f'Datadog said {r.status_code} - check the API key + application key')
    r.raise_for_status()
    rows = [{'name': m.get('name'), 'state': m.get('overall_state'), 'type': m.get('type'),
             'muted': bool((m.get('options') or {}).get('silenced')), 'modified': m.get('modified')}
            for m in r.json()]
    rows.sort(key=lambda m: {'Alert': 0, 'Warn': 1, 'No Data': 2}.get(m['state'], 3))   # trouble first
    lim, mine = row_limit(cfg)
    return rows_out(rows, lim, unit='monitors', mine=mine)


def run_digest(cfg):
    """{"days": 3} - Taskuary's own activity as the data: open work, finished work, pending
    reviews, fresh verdicts, who wrote how often. The Morning digest ships as a report ON
    PURPOSE: the brief lands on the Timeline like any report, its prompt is edited on the
    Reports tab, deleting the source turns it off - and it demonstrates how reports work
    using data every install already has. `store` arrives via resolve_cfg, never persisted."""
    from .digest import gather
    days = int(cfg.get('days') or 3)
    return f'the last {days} days, distilled', gather(cfg['store'], days)


def run_automate(cfg):
    """{"days": 30} - Taskuary's own traffic as the data: what repeats often enough to
    automate, and the concrete policy/report/switch that would kill it. Ships seeded as
    the weekly 'Automation ideas' report; see toil.py. `store` arrives via resolve_cfg."""
    from .toil import gather
    days = int(cfg.get('days') or 30)
    return f'the last {days} days of repeated toil', gather(cfg['store'], days)


def _planned(name):
    def _fail(cfg): raise NotImplementedError(f"connector type '{name}' is on the roadmap - not implemented yet")
    return _fail


def _newest(path: str, by: str = 'newest'):
    """The file a scheduled report should read. A glob is the point, not a convenience: an export
    that lands as sales-2026-08-25.csv has a different name every morning, so a report naming one
    file exactly is a report that works for a day. The NEWEST match is what "the latest export"
    means. Also accepts a folder, and a plain path unchanged."""
    from glob import glob
    from pathlib import Path
    p = Path(path).expanduser()
    if any(ch in str(p) for ch in '*?['):
        files = [h for h in (Path(x) for x in glob(str(p))) if h.is_file()]
        if not files: raise RuntimeError(f'nothing matches {path} - no file to read')
        # mtime is what "the export that just arrived" means, and it is the right default. But a
        # file NAMED by its date is the case where mtime lies: re-copy last month's archive and it
        # becomes the newest file on disk while sales-2026-08-01.csv is plainly not the latest
        # sales. pick='name' takes the highest-sorting name instead, which for an ISO date IS the
        # latest. The headline always says which file was read, so a wrong guess is visible.
        if str(by or 'newest').lower() == 'name': return sorted(files)[-1]
        return max(files, key=lambda h: h.stat().st_mtime)
    if not p.exists(): raise RuntimeError(f'{p} does not exist on this machine')
    return p


def _rows_from_text(text: str, delim: str = None) -> list:
    """A delimited file as dicts, sniffing the separator when it is not given: exports arrive as
    csv, tsv and semicolon-separated depending on who produced them and in what locale."""
    import csv
    sample = text[:4000]
    if not delim:
        try: delim = csv.Sniffer().sniff(sample, delimiters=',;\t|').delimiter
        except csv.Error: delim = ','
    return [dict(r) for r in csv.DictReader(io.StringIO(text), delimiter=delim)]


def run_local_file(cfg):
    """{"path": "C:/exports/*.csv", "tail": 50, "sheet": "Sheet1"} - a file, folder or glob on
    THIS machine. Taskuary already runs on the owner's own computer, so the spreadsheet somebody
    drops in a folder every morning is a report source like any other - and the alternative was
    a WinRM script or nothing.

    csv/tsv/json/jsonl come back as rows; xlsx too where openpyxl is installed. Anything else is
    read as text and the LAST `tail` lines are shown, because a log's news is at the bottom.
    A folder lists what is in it, newest first, which answers "did today's export arrive?"."""
    import json as _json
    p = _newest(cfg['path'], cfg.get('pick'))
    lim, mine = row_limit(cfg)
    if p.is_dir():
        rows = sorted(({'name': f.name, 'bytes': f.stat().st_size,
                        'modified': datetime.fromtimestamp(f.stat().st_mtime).strftime('%Y-%m-%d %H:%M')}
                       for f in p.iterdir() if f.is_file()),
                      key=lambda r: r['modified'], reverse=True)
        return rows_out(rows, lim, unit=f'files in {p.name}', mine=mine)
    suffix = p.suffix.lower()
    if suffix == '.xlsx':
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError('reading .xlsx needs openpyxl - run: pip install openpyxl '
                               '(csv, tsv, json and text need nothing)')
        wb = load_workbook(p, read_only=True, data_only=True)      # data_only: values, not formulae
        ws = wb[cfg['sheet']] if cfg.get('sheet') else wb.active
        it = ws.iter_rows(values_only=True)
        head = [str(h) if h is not None else f'col{i}' for i, h in enumerate(next(it, []) or [])]
        rows = [dict(zip(head, [('' if v is None else v) for v in r])) for r in it]
        wb.close()
        return rows_out(rows, lim, unit=f'rows from {p.name}', mine=mine)
    text = p.read_text(encoding='utf-8', errors='replace')
    if suffix in ('.csv', '.tsv'):
        return rows_out(_rows_from_text(text, '\t' if suffix == '.tsv' else cfg.get('delimiter')),
                        lim, unit=f'rows from {p.name}', mine=mine)
    if suffix == '.jsonl':
        rows = [_json.loads(l) for l in text.splitlines() if l.strip()]
        return rows_out(rows, lim, unit=f'records from {p.name}', mine=mine)
    if suffix == '.json':
        data = _json.loads(text or 'null')
        if cfg.get('path_expr'):
            for k in str(cfg['path_expr']).split('.'):
                if k: data = data[int(k)] if isinstance(data, list) else (data or {}).get(k)
        if isinstance(data, list): return rows_out(data, lim, unit=f'records from {p.name}', mine=mine)
        return f'{p.name}', _json.dumps(data, indent=1, default=str)[:BODY_CHARS]
    lines = text.splitlines()
    try: tail = max(1, int(cfg.get('tail') or 50))
    except (TypeError, ValueError): tail = 50
    shown = lines[-tail:]
    head = f'{p.name} - last {len(shown)} of {len(lines)} lines'
    return head, '\n'.join(shown)[:BODY_CHARS]


def _research(name):
    def run(cfg):
        from . import research
        return getattr(research, f'run_{name}')(cfg)
    run.__doc__ = f'research.run_{name} - see taskuary/research.py'
    return run


def _calendar(cfg):
    from .calendar import run_calendar
    return run_calendar(cfg)

REGISTRY = {'sqlite': run_sqlite, 'mssql': run_mssql, 'database': run_database,
            # the web as a source: plain REST, a key on a card, nothing new in the exe
            'exa': _research('exa'), 'tavily': _research('tavily'),
            'firecrawl': _research('firecrawl'), 'reader': _research('reader'),
            'local_file': run_local_file,
            'aws': run_aws, 's3_object': run_s3, 'cloudwatch_logs': run_cwlogs,
            'azure': run_azure, 'azure_blob': run_azblob, 'azure_logs': run_azlogs,
            'entra_users': run_entra_users, 'entra_groups': run_entra_groups,
            'entra_signins': run_entra_signins, 'entra_licenses': run_entra_licenses,
            'prometheus': run_prometheus, 'datadog': run_datadog,
            'winrm': run_winrm, 'mcp': run_mcp, 'rest': run_rest,
            'intacct': run_intacct, 'intacct_fields': run_intacct_fields,
            'rss': run_rss, 'digest': run_digest, 'automate': run_automate,
            'calendar': _calendar,       # the owner's busy times, off the Outlook (and Google) cards - read-only
            'agent': run_agent,          # the AI itself: a saved skill or a prompt, run by a CLI agent on the schedule
            **{n: _planned(n) for n in PLANNED}}

# Which connector CARD owns each executor type: the s3/cloudwatch types run on the aws
# card's keys, the blob/logs types on the azure card's app - roles and creds resolve there.
CARD_OF = {'s3_object': 'aws', 'cloudwatch_logs': 'aws', 'azure_blob': 'azure', 'azure_logs': 'azure', 'calendar': 'outlook',
           'entra_users': 'azure', 'entra_groups': 'azure', 'entra_signins': 'azure', 'entra_licenses': 'azure',
           'intacct_fields': 'intacct'}

def card_of(t): return CARD_OF.get(t, t)


def mssql_connection(store) -> dict:
    """The SQL Server CONNECTION lives on the mssql connector card (set up once, tested
    there); report configs carry only query/ai_prompt/schedule and inherit it here.
    Per-report overrides still win if present."""
    c = store.get_connector_by_type('mssql', with_secret=True)
    if not c: return {}
    cfg = json.loads(c.get('ConfigJson') or '{}')
    if c.get('Secret'): cfg.setdefault('password', c['Secret'])
    return {k: v for k, v in cfg.items() if v}


def winrm_connection(store) -> dict:
    """Same connection-card pattern as mssql: the host lives on the winrm connector."""
    c = store.get_connector_by_type('winrm')
    cfg = json.loads((c or {}).get('ConfigJson') or '{}')
    return {k: v for k, v in cfg.items() if v}


def _card(store, typ, secret_as):
    c = store.get_connector_by_type(typ, with_secret=True)
    if not c: return {}
    cfg = json.loads(c.get('ConfigJson') or '{}')
    if c.get('Secret'): cfg.setdefault(secret_as, c['Secret'])
    return {k: v for k, v in cfg.items() if v}


def database_connection(store) -> dict:
    """The connection string lives on the 'Any database' card; its write-only secret fills
    the string's {password} placeholder."""
    return _card(store, 'database', 'password')


def aws_connection(store) -> dict:
    return _card(store, 'aws', 'secret_access_key')


def azure_connection(store) -> dict:
    """The Azure card's own app, else the Outlook connector's saved Graph app - one app
    registration can hold Graph permissions AND Azure RBAC roles, so the borrow is real."""
    cfg = _card(store, 'azure', 'client_secret')
    if not (cfg.get('client_id') and cfg.get('client_secret')):
        cfg = {**_card(store, 'outlook', 'client_secret'), **cfg}
    return cfg


def intacct_connection(store) -> dict:
    """Five credentials, of which exactly one is a secret worth hiding: the API USER's
    password. The sender pair identifies the integration and the company id names the tenant -
    neither is a password to this company's books, and burying them write-only would only mean
    nobody can ever check the sender id for a typo."""
    return _card(store, 'intacct', 'user_password')


def prometheus_connection(store) -> dict:
    """base_url (+ optional bearer token as the write-only secret) lives on the card."""
    return _card(store, 'prometheus', 'token')


def datadog_connection(store) -> dict:
    """site + application key on the card; the API key is the write-only secret."""
    return _card(store, 'datadog', 'api_key')


def _apikey_card(typ):
    """A card whose whole configuration is one key: the secret arrives as `api_key`."""
    return lambda store: _card(store, typ, 'api_key')


CONNECTION_OF = {'mssql': mssql_connection, 'winrm': winrm_connection, 'database': database_connection,
                 'exa': _apikey_card('exa'), 'tavily': _apikey_card('tavily'),
                 'firecrawl': _apikey_card('firecrawl'), 'reader': _apikey_card('reader'),
                 'aws': aws_connection, 's3_object': aws_connection, 'cloudwatch_logs': aws_connection,
                 'azure': azure_connection, 'azure_blob': azure_connection, 'azure_logs': azure_connection,
                 'entra_users': azure_connection, 'entra_groups': azure_connection,
                 'entra_signins': azure_connection, 'entra_licenses': azure_connection,
                 'prometheus': prometheus_connection, 'datadog': datadog_connection,
                 'intacct': intacct_connection, 'intacct_fields': intacct_connection}


def resolve_cfg(store, cfg: dict) -> dict:
    if cfg.get('type') in ('digest', 'automate', 'agent', 'calendar'): return {**cfg, 'store': store}   # their data IS the store (the agent's: its profile; the calendar's: the cards)
    conn = CONNECTION_OF.get(cfg.get('type'))
    if conn: return {**conn(store), **{k: v for k, v in cfg.items() if v not in (None, '')}}
    return cfg


AI_SYSTEM = ('You summarize scheduled report data for a busy operator. Follow the operator '
             'instruction exactly. Be concise and concrete: numbers, names, deltas. Plain text only. '
             'The data may be a CAPPED slice of a larger result (the headline says so, and the rows '
             'may be cut mid-way) - never describe a capped or truncated slice as complete, and say '
             'plainly when something the instruction asks about is not present in the rows you got.')

# The rows come back as a spreadsheet and a chart, and the model that just read every row knows
# which column is the measure better than a heuristic hunting for "all numeric" does.
CHART_SYSTEM = ('\n\nThe rows are also turned into a bar chart for the reader. If ONE column is a '
                'measure worth plotting, end your answer with a single line:\n'
                'CHART: <value column> | <label column> | <short chart title>\n'
                'Use the exact column names from the data. Omit the line entirely when the rows are '
                'not worth plotting (no measure, one row, or every value the same) - a chart of '
                'nothing is worse than no chart.')


def run_sources(store, subs: list):
    """Several sources feeding ONE report: each runs on its own connection and query, the
    bodies are stacked under labeled headers, and the AI pass downstream sees all of them
    at once. The same connection can appear twice with different queries. One source
    failing is reported in place - it never takes the whole report down."""
    heads, bodies = [], []
    for i, sub in enumerate(subs, 1):
        t = sub.get('type', 'rest')
        label = (sub.get('label') or '').strip() or f'{t} #{i}'
        try:
            head, body = REGISTRY[t](resolve_cfg(store, dict(sub)))
        except Exception as e:
            head, body = 'FAILED', f'error: {str(e)[:400]}'
            logger.warning(f'report source "{label}" failed: {e}')
        heads.append(f'{label}: {head}')
        bodies.append(f'=== {label} ({head}) ===\n{body}')
    return ' · '.join(heads)[:400], '\n\n'.join(bodies)[:BODY_CHARS]


def report_llm(store, cfg: dict, default_llm):
    """The brain THIS report asked for (cfg['ai_brain'] in /api/brains values, optional
    cfg['ai_model'] override) - a heavier model for the weekly review, the cheap tier for
    pings. Falls back to the caller's default (the triage brain) when unset or broken."""
    if not (cfg.get('ai_brain') or cfg.get('ai_model')): return default_llm
    from .llm import build_llm
    try: return build_llm(store, cfg.get('ai_brain') or None, cfg.get('ai_model') or None) or default_llm
    except Exception as e:
        logger.warning(f'report brain unavailable, using the default: {e}')
        return default_llm


def render_report(store, cfg: dict, llm=None):
    """Run the executor(s), then (optionally) the AI pass: cfg['ai_prompt'] + a configured
    AI connector turn raw rows into the summary that lands on the timeline. The report may
    name its own brain and model (report_llm); `llm` is the default it falls back to."""
    llm = report_llm(store, cfg, llm)
    subs = [s for s in (cfg.get('sources') or []) if s.get('type')]
    if subs:
        head, summary = run_sources(store, subs)
    else:
        cfg = resolve_cfg(store, cfg)
        head, summary = REGISTRY[cfg.get('type', 'rest')](cfg)
    if cfg.get('ai_prompt') and llm:
        try:
            data = summary[:AI_CHARS]
            if len(summary) > AI_CHARS: data += '\n…(data truncated here - later rows were NOT shown to you)'
            charts = str(store.get_settings().get('report_images_enabled') or '1') == '1'
            ai = (llm(AI_SYSTEM + (CHART_SYSTEM if charts else ''),
                      f"Instruction: {cfg['ai_prompt']}\n\nData ({head}):\n{data}",
                      max_tokens=SUMMARY_TOKENS) or '').strip()
            # an empty answer used to file as a bare '--- raw data ---' wall, which reads
            # like the prompt was never run. Say what happened instead.
            if not ai:
                ai = ('(the model returned an empty summary - it may have spent its budget thinking. '
                      'Try a shorter prompt, or a non-reasoning model for report summaries.)')
            return head, f"{ai}\n\n--- raw data ---\n{summary[:4000]}"
        except Exception as e:
            logger.warning(f'AI summary failed for report: {e}')
            return head, f'(AI summary failed: {str(e)[:200]})\n\n{summary}'
    if cfg.get('ai_prompt') and not llm:
        return head, f'(AI prompt set, but no active AI connector - raw data below)\n\n{summary}'
    return head, summary


def _cron_field(spec: str, lo: int, hi: int) -> set:
    """One cron field -> the set of matching values. Supports * , - and /step (numeric only)."""
    out = set()
    for part in str(spec).split(','):
        part, step = (part.split('/', 1) + ['1'])[:2]
        step = int(step)
        if part.strip() in ('*', ''): rng = range(lo, hi + 1)
        elif '-' in part: a, b = part.split('-', 1); rng = range(int(a), int(b) + 1)
        else: v = int(part); rng = range(v, v + 1)
        out.update(x for x in rng if lo <= x <= hi)
        if any(x < lo or x > hi for x in rng): raise ValueError(f'{spec}: out of range {lo}-{hi}')
    return out


def cron_prev(expr: str, now: datetime):
    """The most recent minute matching a 5-field cron (min hour dom month dow) at or before
    `now`, scanning back up to 35 days - None when malformed or nothing matches. Vixie rule:
    dom and dow both restricted means EITHER may match. dow: 0 and 7 are Sunday."""
    try:
        parts = str(expr).split()
        if len(parts) != 5: return None
        mins, hrs, doms, mons, dows = (
            _cron_field(p, lo, hi) for p, (lo, hi) in zip(parts, ((0, 59), (0, 23), (1, 31), (1, 12), (0, 7))))
    except ValueError:
        return None
    dows = {0 if x == 7 else x for x in dows}
    dom_star, dow_star = parts[2] == '*', parts[4] == '*'
    t = now.replace(second=0, microsecond=0)
    for _ in range(35 * 24 * 60):
        cd = (t.weekday() + 1) % 7                        # python Mon=0 -> cron Sun=0
        day_ok = (t.day in doms if dow_star else cd in dows if dom_star
                  else (t.day in doms or cd in dows))
        if t.minute in mins and t.hour in hrs and t.month in mons and day_ok: return t
        t -= timedelta(minutes=1)
    return None


def is_due(cfg: dict, last_polled, startup: bool = False) -> bool:
    # on_startup is local-first scheduling: the app is a window you open, so "when I open
    # it" is a real schedule. Due exactly once per launch - never on the 10-minute auto-sync,
    # and a cron time it would have missed while closed is not its problem.
    if cfg.get('on_startup'): return startup
    now = datetime.now()
    if not last_polled: return True
    try: last = datetime.fromisoformat(str(last_polled)[:19].replace(' ', 'T'))
    except ValueError: return True
    if cfg.get('cron'):
        # due when a scheduled minute passed since the last run. A local app sleeps: a cron
        # slot missed while closed fires on the next poll after reopening, once, not N times.
        prev = cron_prev(cfg['cron'], now)
        if prev is not None: return prev > last
        # malformed expression: fall through to the daily default, never a dead report
    if cfg.get('every_minutes'):
        try: return (now - last).total_seconds() >= float(cfg['every_minutes']) * 60
        except (TypeError, ValueError): pass               # 'every 30' typed as words: daily default
    if cfg.get('daily_at'):
        # tolerant of what people type: '8' and '8:30' both parse; garbage falls back to the
        # daily default instead of an unpack error killing the WHOLE poll thread (it did)
        try:
            hh, mm = (str(cfg['daily_at']).strip() + ':0').split(':')[:2]
            due = now.replace(hour=int(hh), minute=int(mm or 0), second=0, microsecond=0)
            return now >= due and last < due
        except (TypeError, ValueError):
            pass
    return (now - last).total_seconds() >= 24 * 3600


def run_report_source(store, src: dict, llm=None) -> dict:
    """Execute one due report (executor + optional AI pass) and file it on the timeline.
    Errors file visibly too."""
    cfg = json.loads(src.get('ConfigJson') or '{}')
    title = cfg.get('title') or src['Address']
    logger.debug(f'report run: {title} ({cfg.get("type", "rest")}, ai={bool(cfg.get("ai_prompt"))})')
    try:
        head, summary = render_report(store, cfg, llm)
        subject, body = f'{title} — {head}', summary
    except Exception as e:
        subject, body = f'{title} — FAILED', f'Report error: {str(e)[:500]}'
        logger.warning(f'report {src["Address"]} failed: {e}')
    stamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    # the CHART: line is an instruction to Taskuary about what to draw, not prose for the reader:
    # artifacts reads it off `body`, and what gets filed is the summary without it
    from .artifacts import strip_directive
    mid = store.add_message({'TaskId': None, 'ExternalId': f'report:{src["SourceId"]}:{stamp}',
                             'ConversationId': f'report:{src["SourceId"]}', 'Channel': 'report',
                             'SourceName': title, 'Subject': subject, 'FromName': title,
                             'SentAt': stamp, 'BodyText': strip_directive(body),
                             'SourceLink': cfg.get('link'), 'Status': 'feed'})
    store.add_route(mid, None, 'feed', None, 'scheduled report - informational, never a task', [], 'report')
    # the rows are the report: hand back the spreadsheet to open and the chart to look at, not
    # just prose about them. Prose-only reports (an AI summary, a failure) produce neither.
    try:
        from .artifacts import attach_report_output
        made = attach_report_output(store, mid, title, body)
    except Exception as e:
        made = []
        logger.warning(f'report artifacts for {title} failed: {e}')
    store.audit('message', mid, 'report', 'report', 'agent', title)
    # ...and if the report is meant to LEAVE, this is where it turns around. Same run, same row
    # on the timeline - it just travels the other way, and says so.
    if cfg.get('deliver', {}).get('to'):
        try: deliver_report(store, src, cfg, subject, strip_directive(body))
        except Exception as e:
            logger.warning(f'outbound delivery for {title} failed: {e}')
            store.add_route(mid, None, 'feed', None, f'the report ran; sending it out failed: {str(e)[:200]}',
                            [], 'report')
    # the digest report is ALSO what keeps DIGEST.md alive: one run, two homes - the Timeline
    # row you read in the morning, and the doc the Docs tab shows
    if 'digest' in {cfg.get('type'), *(s.get('type') for s in cfg.get('sources') or [])}:
        from .digest import HEADER
        store.save_doc('digest', f'{HEADER}_refreshed {stamp[:16]}_\n\n{strip_directive(body)}\n', 'digest')
    return {'message_id': mid, 'subject': subject, 'files': len(made)}


def deliver_report(store, src: dict, cfg: dict, subject: str, body: str) -> dict:
    """A report that goes OUT: to an address the owner chose, on a channel they picked, either
    after they have read it or straight away.

    `gate` is the whole point and it defaults to 'review'. Everything else in this app holds to
    "nothing sends without you", and a scheduled job that mails your customers on its own would
    be the one place that promise did not hold. Choosing 'auto' is the owner saying, once, that
    THIS report is safe to send unread - not a default they discover afterwards.

    Either way it lands on the timeline as an outbound row, so the funnel shows both directions.
    """
    d = cfg.get('deliver') or {}
    gate = str(d.get('gate') or 'review').lower()
    to = d.get('to') if isinstance(d.get('to'), list) else [x.strip() for x in str(d.get('to') or '').split(',') if x.strip()]
    subj = (d.get('subject') or subject or cfg.get('title') or 'Report').strip()
    stamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    mid = store.add_message({
        'ExternalId': f'out:{src["SourceId"]}:{stamp}', 'ConversationId': f'report:{src["SourceId"]}',
        'Channel': d.get('channel') or 'email', 'SourceName': cfg.get('title') or src['Address'],
        'Subject': subj, 'FromName': 'Taskuary', 'SentAt': stamp, 'BodyText': body,
        'Direction': 'out', 'Status': 'draft' if gate == 'review' else 'sent'})
    who = ', '.join(to) or 'nobody yet'
    if gate == 'review':
        store.add_review({'MessageId': mid, 'Kind': 'outbound', 'Status': 'pending', 'DraftText': body,
                          'Reason': f'{cfg.get("title") or "report"} → {who}. Approve to send it.',
                          'Deliver': json.dumps({'channel': d.get('channel') or 'email', 'to': to, 'subject': subj})})
        store.add_route(mid, None, 'draft', None,
                        f'outbound report waiting for you - approve in Review and it goes to {who}', [], 'report')
        store.audit('message', mid, 'outbound_drafted', 'report', 'agent', {'to': to, 'channel': d.get('channel')})
        return {'gate': 'review', 'message_id': mid, 'to': to}
    from . import outbound
    sent = outbound.send_out(store, d.get('channel') or 'email', to, subj, body)
    store.add_route(mid, None, 'send', None,
                    f'sent automatically to {who} - this report is set to send without review', [], 'report')
    store.audit('message', mid, 'outbound_sent', 'report', 'agent', {'to': to, 'channel': sent.get('channel')})
    return {'gate': 'auto', 'message_id': mid, 'sent': sent}


def run_due_reports(store, startup: bool = False) -> int:
    from .llm import build_llm
    try: llm = build_llm(store)
    except Exception: llm = None
    n = 0
    for src in store.list_sources():
        if src['Channel'] != 'report': continue
        if is_due(json.loads(src.get('ConfigJson') or '{}'), src.get('LastPolledAt'), startup):
            run_report_source(store, src, llm)
            store.touch_source(src['SourceId'])
            n += 1
    return n
